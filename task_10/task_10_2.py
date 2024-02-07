# Задача 10-2
# Дан эксельный файл. На странице находится список сотрудников и их ЗП.
# Надо создать еще одну страницу в этом файле и поместить туда отсортированный
# список фамилий и их ЗП, а в последней строчке поместить слово ИТОГО: и сумму
# всех ЗП. Сортировка по убыванию ЗП сотрудников.
# Например, исходная страница файла:
# Сидоров 100
# Петров 200
# Иванов 300
# Результат:
# Иванов 300
# Петров 200
# Сидоров 100
# ИТОГО: 600

import openpyxl         # импорт библиотеки
from openpyxl import Workbook       # сделал как в уроке, но я этого не понимаю. Мы же уже всю библиотеку импортировали выше
wb = Workbook()

wb = openpyxl.load_workbook('../files/salary.xlsx')
# print(wb.sheetnames)    # имена листов в книге
ws = wb.active
# print(ws.title)       # имя активного листа

sal = {}                # пустой
for a in range(ws.max_row):         #max_row о конца таблички гоним
    sal.update({ws.cell(row=a+1, column=1).value:ws.cell(row=a+1, column=2).value})
#print(sal)
#print(type(sal))

sortedSal = sorted(sal.items(), key=lambda x: -x[1])
#print(sortedSal)

totalSal = sum(value for key, value in sortedSal)
#print(totalSal)

sortedSalary = wb.create_sheet('sortedSalary')


for i, v in sortedSal:
    sortedSalary.append([i, v])

sortedSalary.append(['Total:', totalSal])

wb.save('../files/salary.xlsx')

# Задача 11-2
#
# Дан файл с расширением .txt , содержащий в каждой строчке следующую информацию:
# номер, фамилия, имя, компания, зарплата, разделенные запятыми.
# Создайте Эксельный файл, в который перенесите эту информацию,
# предварительно отсортировав этот список по компании, по фамилии и по имени
# В конце списка добавьте строку: ИТОГО и суммарное значение всех зарплат.

import openpyxl
from openpyxl import Workbook
wb = Workbook()

wb = openpyxl.load_workbook('../files/empData.xlsx')
# print(wb.sheetnames)    # имена листов в книге
ws = wb.active


with open('../files/empData.txt','r', encoding='UTF-8') as f:
    lines = f.readlines()
    #sortLines = sorted(lines, key=lambda x: (x.split(',')[3]))
    # sortLines = sorted(lines, key=lambda x: (x.split(',')[1], x.split(',')[2], x.split(',')[3]))
    sortLines = sorted(lines, key=lambda x: (x.split(',')[3], x.split(',')[1], x.split(',')[2]))
    print(sortLines)



for i in sortLines:
    total = 0
    row = i.strip().split(',')
    print(row)
    total += int(row[4])
    print(total)
    #ws.append(row)

print(total)

wb.save('../files/empData.xlsx')




    # for i in f:
    #     w = ' '.join(sorted(i.split(',') ,key = i[3]))
    #     w = ' '.join((i.split(',')))
    #     #w = ((i.split(',')))
    #     print(w)
    #         # print(w, file=f2, end='\n')


# import openpyxl
# with open('test.txt') as f:
#     lst = []
#     for i in f:
#         lst.append(i.strip().split(','))
# wb = openpyxl.load_workbook('test.xlsx')
# if 'Sheet' not in wb.sheetnames:
#     wb.create_sheet('Sheet')
# ws=wb['Sheet']
# for i in range(1, ws.max_row+1):
#     for j in range(1, ws.max_column+1):
#         ws.cell(i, j).value = None
# lst_sorted =  sorted(lst, key=lambda x: (x[3],[1],[2]))
# total = 0
# for i, j in enumerate (lst_sorted, 1):
#     for k in range(5):
#         ws.cell(i, k+1).value = j[k]
#     total += int([4])
# ws.cell(i+1, 2).value = 'Итого:'
# ws.cell(i+1, 5).value = total
# wb.save('test.xlsx')
import openpyxl
# from openpyxl import Workbook
# wb = Workbook()
# wb.save('../files/workBook_1.xlsx')

wb = openpyxl.load_workbook('../files/workBook_1.xlsx')
# print(wb.sheetnames)
# #
ws = wb.active
# print(ws.title)

# wb.create_sheet('new')
# print(wb.sheetnames)
# wb.create_sheet('new')
# print(wb.sheetnames)
# wb.remove('new')
# print(wb.sheetnames)

# ws['A1'].value = 123
# print(ws['A1'].value)
# ws['B4'].value = 'ячейка Б2'
# # print(ws['B2'].value)
# print(ws.max_row, ws.max_column)

# for i in range( ws.max_row):
#     for j in range ( ws.max_column):
#         print(i+1,j+1, ws.cell(row = i+1, column=j+1).value)


for i in range(1, ws.max_column):
    print(i, ws.cell(row = 1, column=2).value)

wb.save('../files/workBook_1.xlsx')
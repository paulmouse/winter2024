# Дан исходный эксельный файл со списком людей и их ЗП.
# Следует создать еще одну страницу со статистическими данными исходного списка.
# Например, исходная страница:
# Сидоров 100
# Петров 200
# Иванов 300
# Федоров 1000
# Результат:
# Максимальное значение 1000
# Минимальное значение 100
# Сумма 1600
# Среднее арифметическое 400
# # Медиана 250

import openpyxl         # импорт библиотеки
from openpyxl import Workbook       # сделал как в уроке, но я этого не понимаю. Мы же уже всю библиотеку импортировали выше
wb = Workbook()

wb = openpyxl.load_workbook('../files/salary_data.xlsx')
# print(wb.sheetnames)    # имена листов в книге
ws = wb.active
# print(ws.title)       # имя активного листа

sal = {}                # пустой
for a in range(ws.max_row):         #max_row о конца таблички гоним
    sal.update({ws.cell(row=a+1, column=1).value:ws.cell(row=a+1, column=2).value})
#print(sal)
#print(type(sal))

sortedSal = sorted(sal.items(), key=lambda x: -x[1])
#print(sortedSal)

totalSal = sum(value for key, value in sortedSal)
#print(totalSal)
maxSal = max(value for key, value in sortedSal)
#print(maxSal)
minSal = min(value for key, value in sortedSal)
#print(minSal)
aveSal = totalSal / len(sortedSal)
#print(aveSal)

salaryData = wb.create_sheet('salaryData')


# for i, v in sortedSal:
#     sortedSalary.append([i, v])

salaryData.append(['Total:', totalSal])
salaryData.append(['Max:', maxSal])
salaryData.append(['Min:', minSal])
salaryData.append(['Average:', aveSal])


wb.save('../files/salary_data.xlsx')